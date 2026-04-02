
"""File parsers for both attendance source formats."""

from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from pathlib import Path
import re
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
import xlrd

from .models import EventRecord

DATE_TEXT_PATTERN = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
HEADER_NORMALIZE_PATTERN = re.compile(r"[^a-zа-я0-9]+")


class UnsupportedWorkbookFormatError(RuntimeError):
    """Raised when a workbook cannot be interpreted as expected input type."""


@dataclass
class ParsingSummary:
    """Collects parser statistics for CLI output."""

    processed_files: List[Path] = field(default_factory=list)
    skipped_files: List[Tuple[Path, str]] = field(default_factory=list)
    total_events: int = 0


def parse_directory(input_dir: Path) -> Tuple[List[EventRecord], ParsingSummary]:
    """Parse all supported files in the input directory and return event records."""
    summary = ParsingSummary()
    all_events: List[EventRecord] = []

    for file_path in sorted(input_dir.iterdir()):
        if not file_path.is_file():
            continue

        # Временные lock-файлы Excel (~$...) не являются книгами
        if file_path.name.startswith("~$"):
            continue

        # Не разбираем файлы режима работы/расписания как посещаемость
        if file_path.name in WORK_MODE_FILENAMES:
            continue

        suffix = file_path.suffix.lower()
        if suffix == ".xls":
            try:
                events = parse_first_type_events(file_path)
            except Exception as exc:
                summary.skipped_files.append((file_path, str(exc)))
                continue
            all_events.extend(events)
            summary.processed_files.append(file_path)
            summary.total_events += len(events)
            continue

        if suffix == ".xlsx":
            try:
                events = parse_second_type_events(file_path)
            except UnsupportedWorkbookFormatError as exc:
                summary.skipped_files.append((file_path, str(exc)))
                continue
            except Exception as exc:
                summary.skipped_files.append((file_path, str(exc)))
                continue
            all_events.extend(events)
            summary.processed_files.append(file_path)
            summary.total_events += len(events)
            continue

    return all_events, summary


def parse_first_type_events(file_path: Path) -> List[EventRecord]:
    """Parse first format workbook (`*.xls`) and return employee event records."""
    employee_name = _extract_employee_name_from_stem(file_path.stem)

    try:
        events = _parse_first_type_with_xlrd(file_path, employee_name)
    except Exception:
        events = _parse_first_type_with_openpyxl_fallback(file_path, employee_name)

    if not events:
        raise UnsupportedWorkbookFormatError(
            f"File {file_path.name} does not contain first-format events."
        )
    return events


def parse_second_type_events(file_path: Path) -> List[EventRecord]:
    """Parse second format workbook (`*.xlsx`) and return employee event records."""
    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise UnsupportedWorkbookFormatError(
            f"File {file_path.name} has no header row."
        )

    indexes = _detect_second_type_indexes(header_row)
    if indexes["last_name"] is None or indexes["first_name"] is None:
        raise UnsupportedWorkbookFormatError(
            f"File {file_path.name} is not a second-format workbook "
            "(missing LastName/FirstName columns)."
        )
    if not indexes["time_columns"]:
        raise UnsupportedWorkbookFormatError(
            f"File {file_path.name} is not a second-format workbook "
            "(missing Event time columns)."
        )

    records: List[EventRecord] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        fio = _compose_fio(
            row[indexes["last_name"]],
            row[indexes["first_name"]],
            row[indexes["middle_name"]] if indexes["middle_name"] is not None else None,
        )
        if not fio:
            continue

        # Optional direction column (EntryExit with values "Вход"/"Выход")
        is_entry: Optional[bool] = None
        entry_exit_index = indexes.get("entry_exit")
        if entry_exit_index is not None:
            raw_direction = row[entry_exit_index]
            if isinstance(raw_direction, str):
                direction_clean = raw_direction.strip().lower()
                if direction_clean == "вход":
                    is_entry = True
                elif direction_clean == "выход":
                    is_entry = False

        for time_index in indexes["time_columns"]:
            timestamp = _parse_timestamp(row[time_index])
            if timestamp is None:
                continue
            records.append(
                EventRecord(
                    employee_name=fio,
                    occurred_at=timestamp,
                    is_entry=is_entry,
                )
            )

    if not records:
        raise UnsupportedWorkbookFormatError(
            f"File {file_path.name} has matching headers but no parsable events."
        )
    return records


def _parse_first_type_with_xlrd(file_path: Path, employee_name: str) -> List[EventRecord]:
    """Parse first-format workbook with xlrd for legacy BIFF `.xls` files."""
    workbook = xlrd.open_workbook(file_path.as_posix())
    sheet = workbook.sheet_by_index(0)

    records: List[EventRecord] = []
    current_date: Optional[date] = None

    for row_idx in range(sheet.nrows):
        cell = sheet.cell(row_idx, 0)
        raw_value = cell.value
        cell_type = cell.ctype

        parsed_date = _parse_date_from_value(raw_value)
        if parsed_date is not None:
            current_date = parsed_date
            continue

        if cell_type == xlrd.XL_CELL_DATE:
            parsed_datetime = xlrd.xldate_as_datetime(raw_value, workbook.datemode)
            if parsed_datetime.date().year >= 1900 and parsed_datetime.time() == time(0, 0):
                current_date = parsed_datetime.date()
                continue
            if current_date is not None:
                records.append(
                    EventRecord(
                        employee_name=employee_name,
                        occurred_at=datetime.combine(
                            current_date, parsed_datetime.time()
                        ),
                    )
                )
            continue

        parsed_time = _parse_time_from_value(raw_value)
        if parsed_time is not None and current_date is not None:
            records.append(
                EventRecord(
                    employee_name=employee_name,
                    occurred_at=datetime.combine(current_date, parsed_time),
                )
            )

    return records


def _parse_first_type_with_openpyxl_fallback(
    file_path: Path, employee_name: str
) -> List[EventRecord]:
    """Parse first-format workbook via openpyxl when extension/content mismatch occurs."""
    with file_path.open("rb") as workbook_stream:
        workbook = openpyxl.load_workbook(
            workbook_stream, data_only=True, read_only=True
        )
    sheet = workbook.active

    records: List[EventRecord] = []
    current_date: Optional[date] = None

    for (raw_value,) in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
        parsed_date = _parse_date_from_value(raw_value)
        if parsed_date is not None:
            current_date = parsed_date
            continue

        if isinstance(raw_value, datetime):
            if current_date is None:
                current_date = raw_value.date()
            records.append(
                EventRecord(
                    employee_name=employee_name,
                    occurred_at=datetime.combine(current_date, raw_value.time()),
                )
            )
            continue

        parsed_time = _parse_time_from_value(raw_value)
        if parsed_time is not None and current_date is not None:
            records.append(
                EventRecord(
                    employee_name=employee_name,
                    occurred_at=datetime.combine(current_date, parsed_time),
                )
            )

    return records


def _detect_second_type_indexes(header_row: Sequence[object]) -> Dict[str, object]:
    """Detect required column indexes in second-format workbook headers."""
    index_map: Dict[str, object] = {
        "last_name": None,
        "first_name": None,
        "middle_name": None,
        "entry_exit": None,
        "time_columns": [],
    }

    for idx, header in enumerate(header_row):
        normalized = _normalize_header(header)
        if not normalized:
            continue

        if index_map["last_name"] is None and (
            "lastname" in normalized or "фамилия" in normalized
        ):
            index_map["last_name"] = idx

        if index_map["first_name"] is None and (
            "firstname" in normalized
            or "имяпользователя" in normalized
            or normalized == "имя"
        ):
            index_map["first_name"] = idx

        if index_map["middle_name"] is None and (
            "middlename" in normalized or "отчество" in normalized
        ):
            index_map["middle_name"] = idx

        is_kb_time = (
            "receipttime" in normalized
            or "времяпкб" in normalized
            or "времяпокб" in normalized
            or ("event" in normalized and "кб" in normalized)
        )
        is_event_time = (
            normalized == "eventtime"
            or ("eventtime" in normalized and "receipt" not in normalized)
            or ("event" in normalized and "time" in normalized)
        )

        if is_kb_time or is_event_time:
            index_map["time_columns"].append(idx)

        # Entry/Exit direction (e.g. "EntryExit", "Entry Exit")
        if index_map["entry_exit"] is None and ("entryexit" in normalized):
            index_map["entry_exit"] = idx

    index_map["time_columns"] = sorted(set(index_map["time_columns"]))
    return index_map


def _compose_fio(last_name: object, first_name: object, middle_name: object) -> str:
    """Build canonical `LastName FirstName MiddleName` string from row values."""
    parts = [str(part).strip() for part in (last_name, first_name, middle_name) if part]
    parts = [part for part in parts if part]
    return " ".join(parts)


def _parse_timestamp(raw_value: object) -> Optional[datetime]:
    """Parse mixed timestamp value to naive datetime."""
    if raw_value is None:
        return None

    if isinstance(raw_value, datetime):
        if raw_value.tzinfo is not None:
            return raw_value.replace(tzinfo=None)
        return raw_value

    if isinstance(raw_value, str):
        cleaned = raw_value.strip()
        if not cleaned:
            return None

        if cleaned.endswith("Z"):
            cleaned = cleaned[:-1] + "+00:00"

        try:
            parsed = datetime.fromisoformat(cleaned)
            return parsed.replace(tzinfo=None)
        except ValueError:
            pass

        known_formats = (
            "%Y-%m-%d %H:%M:%S.%f",
            "%Y-%m-%d %H:%M:%S",
            "%d.%m.%Y %H:%M:%S",
            "%d.%m.%Y %H:%M",
        )
        for date_format in known_formats:
            try:
                return datetime.strptime(cleaned, date_format)
            except ValueError:
                continue

    return None


def _parse_date_from_value(raw_value: object) -> Optional[date]:
    """Try parsing a date from mixed workbook value representations."""
    if isinstance(raw_value, datetime):
        return raw_value.date()
    if isinstance(raw_value, date):
        return raw_value
    if not isinstance(raw_value, str):
        return None

    cleaned = raw_value.strip()
    if not cleaned or not DATE_TEXT_PATTERN.match(cleaned):
        return None
    return datetime.strptime(cleaned, "%d.%m.%Y").date()


def _parse_time_from_value(raw_value: object) -> Optional[time]:
    """Try parsing a time from numeric, `datetime` or text workbook values."""
    if isinstance(raw_value, datetime):
        return raw_value.time()
    if isinstance(raw_value, time):
        return raw_value

    if isinstance(raw_value, (int, float)) and 0 <= float(raw_value) < 1:
        seconds = int(round(float(raw_value) * 24 * 60 * 60)) % (24 * 60 * 60)
        base = datetime.min + timedelta(seconds=seconds)
        return base.time()

    if isinstance(raw_value, str):
        cleaned = raw_value.strip()
        if not cleaned:
            return None
        for time_format in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(cleaned, time_format).time()
            except ValueError:
                continue

    return None


def _normalize_header(raw_header: object) -> str:
    """Normalize a header value for resilient keyword matching."""
    text = str(raw_header or "").strip().lower()
    return HEADER_NORMALIZE_PATTERN.sub("", text)

def _extract_employee_name_from_stem(stem: str) -> str:
    """Extract employee display name (FIO only) from filename stem."""
    # Удаляем внешние одинарные кавычки, если есть
    cleaned = stem.strip("'\"")
    # Ищем последовательность из 2-4 русских слов (Фамилия Имя [Отчество]) до _ или конца
    match = re.match(r'^([А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+(?:\s+[А-ЯЁ][а-яё]+)?)', cleaned)
    if match:
        return match.group(1).strip()
    # Fallback: оригинальная логика
    fallback = re.sub(r"[_\s]+", " ", cleaned).strip()
    return fallback or cleaned


WORK_MODE_FILENAMES = ("Режим работы.xlsx", "Режимы работы.xlsx", "Расписание работы.xlsx")


@dataclass(frozen=True)
class WorkModeInfo:
    mode: str
    start_time: Optional[time]
    end_time: Optional[time]
    # Business rule flags: were values explicitly stored as 00:00:00 in source?
    start_was_zero: bool = False
    end_was_zero: bool = False


def _load_work_mode_info_from_path(path: Path) -> Dict[str, WorkModeInfo]:
    """Parse a single workbook path into ФИО -> WorkModeInfo."""
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        workbook.close()
        return {}

    fio_idx: Optional[int] = None
    mode_idx: Optional[int] = None
    start_idx: Optional[int] = None
    end_idx: Optional[int] = None

    for idx, header in enumerate(header_row):
        norm = _normalize_header(header)
        if norm in ("фио", "фиосотрудника"):
            fio_idx = idx
        if norm in ("режимработы", "режимыработы"):
            mode_idx = idx
        if norm == "началонормрабочврем":
            start_idx = idx
        if norm in ("конецнормрабчвр", "конецнормрабочвр"):
            end_idx = idx

    if fio_idx is None:
        workbook.close()
        return {}

    result: Dict[str, WorkModeInfo] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) <= fio_idx:
            continue
        fio_val = row[fio_idx]
        if fio_val is None or not str(fio_val).strip():
            continue
        fio = str(fio_val).strip()

        mode_val = row[mode_idx] if (mode_idx is not None and len(row) > mode_idx) else None
        start_val = row[start_idx] if (start_idx is not None and len(row) > start_idx) else None
        end_val = row[end_idx] if (end_idx is not None and len(row) > end_idx) else None

        mode = str(mode_val).strip() if mode_val is not None else ""
        start_time = _parse_time_from_value(start_val)
        end_time = _parse_time_from_value(end_val)

        # Some schedule files use 00:00:00 to indicate "unspecified".
        # Business rule: treat missing start as 09:00 and missing end as 18:00,
        # and remember this substitution for coloring.
        start_was_zero = start_time == time(0, 0)
        end_was_zero = end_time == time(0, 0)

        if start_was_zero:
            start_time = time(9, 0)
        if end_was_zero:
            end_time = time(18, 0)

        result[fio] = WorkModeInfo(
            mode=mode,
            start_time=start_time,
            end_time=end_time,
            start_was_zero=start_was_zero,
            end_was_zero=end_was_zero,
        )

    workbook.close()
    return result


def _merge_work_mode_info(
    existing: WorkModeInfo, incoming: WorkModeInfo
) -> WorkModeInfo:
    """Merge rows from multiple workbooks: non-empty incoming fields override."""
    new_mode = incoming.mode.strip() if incoming.mode else existing.mode
    return WorkModeInfo(
        mode=new_mode,
        start_time=incoming.start_time
        if incoming.start_time is not None
        else existing.start_time,
        end_time=incoming.end_time if incoming.end_time is not None else existing.end_time,
        start_was_zero=incoming.start_was_zero
        if incoming.start_time is not None
        else existing.start_was_zero,
        end_was_zero=incoming.end_was_zero
        if incoming.end_time is not None
        else existing.end_was_zero,
    )


def load_work_mode_info(input_dir: Path) -> Optional[Dict[str, WorkModeInfo]]:
    """Load employee work mode and normative start/end times from work mode workbooks.

    Reads every file listed in `WORK_MODE_FILENAMES` that exists under `input_dir`,
    in order, and merges rows by ФИО. Later files override missing fields from
    earlier files (so e.g. `Расписание работы.xlsx` can add times without clearing
    режим from `Режимы работы.xlsx`).

    Expected columns (best-effort):
    - ФИО / Ф.И.О. сотрудника
    - режим работы (optional in schedule-only files)
    - НачалоНормРабочВрем
    - КонецНормРабчВр / КонецНормРабочВр

    Returns None if no workbook is present; otherwise merged dict.
    """
    existing_in_order = [n for n in WORK_MODE_FILENAMES if (input_dir / n).is_file()]
    if not existing_in_order:
        return None

    merged: Dict[str, WorkModeInfo] = {}

    for name in WORK_MODE_FILENAMES:
        candidate = input_dir / name
        if not candidate.is_file():
            continue
        partial = _load_work_mode_info_from_path(candidate)

        for fio, incoming in partial.items():
            if fio in merged:
                merged[fio] = _merge_work_mode_info(merged[fio], incoming)
            else:
                merged[fio] = incoming

    return merged


def load_work_mode_mapping(input_dir: Path) -> Optional[Dict[str, str]]:
    """Load ФИО -> режим работы from 'Режим работы.xlsx' or 'Режимы работы.xlsx' in input_dir.

    File must have columns 'ФИО' and 'режим работы'. Returns None if neither file
    is present; otherwise a dict (later row wins on duplicate FIO).
    """
    info = load_work_mode_info(input_dir)
    if info is None:
        return None
    return {fio: row.mode for fio, row in info.items()}