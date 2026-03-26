"""Excel report generation using formulas for all derived values."""

from datetime import date, time
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import DayBounds, EmployeeCalendar

HEADER_FILL = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")
LIGHT_RED_FILL = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
DARK_RED_FILL = PatternFill(fill_type="solid", start_color="C45050", end_color="C45050")
LIGHT_GREEN_FILL = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
DARK_GREEN_FILL = PatternFill(fill_type="solid", start_color="2D7D2F", end_color="2D7D2F")
WEEKEND_GREY_FILL = PatternFill(
    fill_type="solid", start_color="EDEDED", end_color="EDEDED"
)
YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
BLACK_FONT = Font(color="000000")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
THICK_SIDE = Side(style="thick")

WORK_MODE_FILE_MISSING = "Не найден файл режима работы"
WORK_MODE_EMPLOYEE_NOT_FOUND = "Информация по данному сотруднику не найдена"


def write_report(
    output_path: Path,
    calendar: EmployeeCalendar,
    default_official_time: time = time(9, 0),
    work_mode_by_fio: Optional[Dict[str, str]] = None,
) -> None:
    default_leave_time: time = time(18, 0)

    """Create report workbook from aggregated employee daily attendance bounds."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчет"

    all_dates = _collect_sorted_dates(calendar)
    _write_header(sheet, all_dates)
    _write_body(
        sheet,
        calendar,
        all_dates,
        default_official_time,
        default_leave_time,
        work_mode_by_fio,
    )
    _apply_layout(sheet, all_dates)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _collect_sorted_dates(calendar: EmployeeCalendar) -> List[date]:
    """Collect and sort all dates present in the calendar."""
    day_set = set()
    for employee_days in calendar.values():
        day_set.update(employee_days.keys())
    return sorted(day_set)


def _write_header(sheet: Worksheet, all_dates: List[date]) -> None:
    """Write the first row with fixed headers, date columns and average columns."""
    sheet.cell(row=1, column=1, value="ФИО")
    sheet.cell(row=1, column=2, value="Начало рабочего дня")
    sheet.cell(row=1, column=3, value="Конец рабочего дня")
    sheet.cell(row=1, column=4, value="Продолжительность работы")
    sheet.cell(row=1, column=5, value="Режим работы")
    sheet.cell(row=1, column=6, value="Показатель")

    first_date_column = 7
    for date_index, day_value in enumerate(all_dates):
        column_index = first_date_column + date_index
        cell = sheet.cell(row=1, column=column_index, value=day_value)
        cell.number_format = "dd.mm.yyyy"

        # Second header row under each date column: "Будний"/"Выходной".
        # Weekend is Saturday+Sunday.
        header_label = "Выходной" if day_value.weekday() >= 5 else "Будний"
        sheet.cell(row=2, column=column_index, value=header_label)

        # Third header row: "Сокращенный" for the day immediately before
        # Saturday in the sorted `all_dates` list.
        short_label = (
            "Сокращенный"
            if date_index + 1 < len(all_dates) and all_dates[date_index + 1].weekday() == 5
            else ""
        )
        if short_label:
            sheet.cell(row=3, column=column_index, value=short_label)

    average_headers = (
        "Среднее время прихода",
        "Среднее время ухода",
        "Среднее время работы",
        "Среднее отклонение прихода",
        "Среднее время переработок",
        "Среднее отсутствие в течение дня",
        "Средняя длительность факт (с учетом отсутствия в течение дня)",
        "Переработка факт (с учетом отсутствия в течение дня)",
    )
    first_average_column = first_date_column + len(all_dates)
    for offset, label in enumerate(average_headers):
        sheet.cell(row=1, column=first_average_column + offset, value=label)


def _write_body(
    sheet: Worksheet,
    calendar: EmployeeCalendar,
    all_dates: List[date],
    default_official_time: time,
    default_leave_time: time,
    work_mode_by_fio: Optional[Dict[str, str]] = None,
) -> None:
    """Write employee rows with values and formulas."""
    first_data_column = 7
    first_average_column = first_data_column + len(all_dates)

    if work_mode_by_fio is None:
        work_mode_label = WORK_MODE_FILE_MISSING
    else:
        work_mode_label = None  # resolve per employee

    for employee_index, employee_name in enumerate(sorted(calendar)):
        arrival_row = 4 + employee_index * 8
        leave_row = arrival_row + 1
        work_row = arrival_row + 2
        delta_row = arrival_row + 3
        overtime_row = arrival_row + 4
        absence_row = arrival_row + 5
        work_minus_absence_row = arrival_row + 6
        overtime_minus_absence_row = arrival_row + 7

        sheet.cell(row=arrival_row, column=1, value=employee_name)

        if work_mode_label is not None:
            sheet.cell(row=arrival_row, column=5, value=work_mode_label)
        else:
            mode = work_mode_by_fio.get(
                employee_name, WORK_MODE_EMPLOYEE_NOT_FOUND
            )
            sheet.cell(row=arrival_row, column=5, value=mode)

        sheet.cell(row=arrival_row, column=2, value=default_official_time)
        sheet.cell(row=arrival_row, column=2).number_format = "hh:mm"

        sheet.cell(row=arrival_row, column=4, value=time(9, 0))
        sheet.cell(row=arrival_row, column=4).number_format = "hh:mm"

        # Конец рабочего дня = начало рабочего дня + продолжительность работы
        end_of_day_formula = f"=B{arrival_row}+D{arrival_row}"
        sheet.cell(row=arrival_row, column=3, value=end_of_day_formula)
        sheet.cell(row=arrival_row, column=3).number_format = "hh:mm"

        sheet.cell(row=arrival_row, column=6, value="Время прихода")
        sheet.cell(row=leave_row, column=6, value="Время ухода")
        sheet.cell(row=work_row, column=6, value="Длительность факт")
        sheet.cell(row=delta_row, column=6, value="Отклонение по времени прихода")
        sheet.cell(row=overtime_row, column=6, value="Переработка")
        sheet.cell(row=absence_row, column=6, value="Отсутствие в течение дня")
        sheet.cell(
            row=work_minus_absence_row,
            column=6,
            value="Длительность факт (с учетом отсутствия в течение дня)",
        )
        sheet.cell(
            row=overtime_minus_absence_row,
            column=6,
            value="Переработка факт (с учетом отсутствия в течение дня)",
        )

        employee_days: Dict[date, DayBounds] = calendar[employee_name]
        for day_offset, day_value in enumerate(all_dates):
            column_index = first_data_column + day_offset
            column_letter = get_column_letter(column_index)
            day_type_cell = f"{column_letter}$2"
            short_day_cell = f"{column_letter}$3"
            plan_duration_expr = (
                f'IF({short_day_cell}="Сокращенный",$D${arrival_row}-TIME(1,15,0),'
                f'$D${arrival_row})'
            )

            day_bounds = employee_days.get(day_value)
            if day_bounds is not None:
                arrival_cell = sheet.cell(
                    row=arrival_row,
                    column=column_index,
                    value=day_bounds.arrival_time,
                )
                leave_cell = sheet.cell(
                    row=leave_row,
                    column=column_index,
                    value=day_bounds.departure_time,
                )
                absence_cell = sheet.cell(
                    row=absence_row,
                    column=column_index,
                    value=day_bounds.absence_duration,
                )
                arrival_cell.number_format = "hh:mm"
                leave_cell.number_format = "hh:mm"
                if day_value.weekday() < 5 and day_bounds.departure_time_fallback:
                    leave_cell.fill = YELLOW_FILL
                absence_cell.number_format = "[h]:mm"

            work_formula = (
                f'=IF(OR({column_letter}{arrival_row}="",{column_letter}{leave_row}=""),"",'
                f"{column_letter}{leave_row}-{column_letter}{arrival_row})"
            )
            delta_formula_workday = (
                f'=IF({column_letter}{arrival_row}="","",'
                f'IF({column_letter}{arrival_row}>=$B${arrival_row},'
                f'TEXT({column_letter}{arrival_row}-$B${arrival_row},"ч:мм"),'
                f'TEXT($B${arrival_row}-{column_letter}{arrival_row},"-ч:мм")))'
            )
            overtime_formula_workday = (
                f'=IF({column_letter}{work_row}="","",'
                f'IF({column_letter}{work_row}>={plan_duration_expr},'
                f'TEXT({column_letter}{work_row}-({plan_duration_expr}),"ч:мм"),'
                f'TEXT(({plan_duration_expr})-{column_letter}{work_row},"-ч:мм")))'
            )
            work_minus_absence_formula = (
                f'=IF(OR({column_letter}{work_row}="",{column_letter}{absence_row}=""),"",'
                f"{column_letter}{work_row}-{column_letter}{absence_row})"
            )
            # Переработка факт (с учетом отсутствия): переработка считается
            # от фактической длительности с учетом отсутствия, а не как
            # разность текстовой переработки и отсутствия.
            overtime_minus_absence_formula_workday = (
                f'=IF({column_letter}{work_minus_absence_row}="","",'
                f'IF({column_letter}{work_minus_absence_row}>={plan_duration_expr},'
                f'TEXT({column_letter}{work_minus_absence_row}-({plan_duration_expr}),"ч:мм"),'
                f'TEXT(({plan_duration_expr})-{column_letter}{work_minus_absence_row},"-ч:мм")))'
            )

            delta_formula = f'=IF({day_type_cell}="Выходной","—",{delta_formula_workday[1:]})'
            overtime_formula = (
                f'=IF({day_type_cell}="Выходной","—",{overtime_formula_workday[1:]})'
            )
            overtime_minus_absence_formula = (
                f'=IF({day_type_cell}="Выходной","—",{overtime_minus_absence_formula_workday[1:]})'
            )

            work_cell = sheet.cell(
                row=work_row, column=column_index, value=work_formula
            )
            delta_cell = sheet.cell(
                row=delta_row, column=column_index, value=delta_formula
            )
            work_minus_absence_cell = sheet.cell(
                row=work_minus_absence_row,
                column=column_index,
                value=work_minus_absence_formula,
            )
            overtime_minus_absence_cell = sheet.cell(
                row=overtime_minus_absence_row,
                column=column_index,
                value=overtime_minus_absence_formula,
            )
            work_cell.number_format = "[h]:mm"
            delta_cell.number_format = "@"
            sheet.cell(
                row=overtime_row, column=column_index, value=overtime_formula
            ).number_format = "[h]:mm"
            work_minus_absence_cell.number_format = "[h]:mm"
            overtime_minus_absence_cell.number_format = "[h]:mm"

        if all_dates:
            start_column_letter = get_column_letter(first_data_column)
            end_column_letter = get_column_letter(first_data_column + len(all_dates) - 1)
            day_type_range = f"{start_column_letter}$2:{end_column_letter}$2"
            short_type_range = f"{start_column_letter}$3:{end_column_letter}$3"

            avg_arrival_formula = (
                f'=IFERROR(AVERAGEIF({day_type_range},"Будний",'
                f"{start_column_letter}{arrival_row}:{end_column_letter}{arrival_row}),\"\")"
            )
            avg_leave_formula = (
                f'=IFERROR(AVERAGEIF({day_type_range},"Будний",'
                f"{start_column_letter}{leave_row}:{end_column_letter}{leave_row}),\"\")"
            )
            avg_work_formula = (
                f'=IFERROR(AVERAGEIF({day_type_range},"Будний",'
                f"{start_column_letter}{work_row}:{end_column_letter}{work_row}),\"\")"
            )
            avg_delta_formula = (
                f'=IFERROR(IF(AVERAGEIF({day_type_range},"Будний",'
                f"{start_column_letter}{arrival_row}:{end_column_letter}{arrival_row})"
                f'>=$B${arrival_row},'
                f'TEXT(AVERAGEIF({day_type_range},"Будний",'
                f"{start_column_letter}{arrival_row}:{end_column_letter}{arrival_row})"
                f'-$B${arrival_row},"ч:мм"),'
                f'TEXT($B${arrival_row}-AVERAGEIF({day_type_range},"Будний",'
                f"{start_column_letter}{arrival_row}:{end_column_letter}{arrival_row}),\"-ч:мм\")),\"\")"
            )
            avg_plan_expr = (
                f'$D${arrival_row}-TIME(1,15,0)*('
                f'COUNTIFS({day_type_range},"Будний",{short_type_range},"Сокращенный")'
                f'/COUNTIF({day_type_range},"Будний")'
                f')'
            )
            avg_overtime_formula = (
                f'=IFERROR(IF(AVERAGEIF({day_type_range},"Будний",'
                f"{start_column_letter}{work_row}:{end_column_letter}{work_row})"
                f'>={avg_plan_expr},'
                f'TEXT(AVERAGEIF({day_type_range},"Будний",'
                f"{start_column_letter}{work_row}:{end_column_letter}{work_row})"
                f'-({avg_plan_expr}),"ч:мм"),'
                f'TEXT(({avg_plan_expr})-AVERAGEIF({day_type_range},"Будний",'
                f"{start_column_letter}{work_row}:{end_column_letter}{work_row}),\"-ч:мм\")),\"\")"
            )
            avg_absence_formula = (
                f'=IFERROR(AVERAGEIF({day_type_range},"Будний",'
                f"{start_column_letter}{absence_row}:{end_column_letter}{absence_row}),\"\")"
            )
            avg_work_minus_absence_formula = (
                f'=IFERROR(AVERAGEIF({day_type_range},"Будний",'
                f"{start_column_letter}{work_minus_absence_row}:{end_column_letter}{work_minus_absence_row}),\"\")"
            )
            # Средняя «переработка факт (с учетом отсутствия)» считается
            # по среднему значению длительности с учетом отсутствия.
            avg_overtime_minus_absence_formula = (
                f'=IFERROR(IF(AVERAGEIF({day_type_range},"Будний",'
                f"{start_column_letter}{work_minus_absence_row}:{end_column_letter}{work_minus_absence_row})"
                f'>={avg_plan_expr},'
                f'TEXT(AVERAGEIF({day_type_range},"Будний",'
                f"{start_column_letter}{work_minus_absence_row}:{end_column_letter}{work_minus_absence_row})"
                f'-({avg_plan_expr}),"ч:мм"),'
                f'TEXT(({avg_plan_expr})-AVERAGEIF({day_type_range},"Будний",'
                f"{start_column_letter}{work_minus_absence_row}:{end_column_letter}{work_minus_absence_row}),\"-ч:мм\")),\"\")"
            )

            avg_arrival_cell = sheet.cell(
                row=arrival_row, column=first_average_column, value=avg_arrival_formula
            )
            avg_leave_cell = sheet.cell(
                row=leave_row, column=first_average_column + 1, value=avg_leave_formula
            )
            avg_work_cell = sheet.cell(
                row=work_row, column=first_average_column + 2, value=avg_work_formula
            )
            avg_delta_cell = sheet.cell(
                row=delta_row, column=first_average_column + 3, value=avg_delta_formula
            )
            avg_overtime_cell = sheet.cell(
                row=overtime_row,
                column=first_average_column + 4,
                value=avg_overtime_formula,
            )
            avg_absence_cell = sheet.cell(
                row=absence_row,
                column=first_average_column + 5,
                value=avg_absence_formula,
            )
            avg_work_minus_absence_cell = sheet.cell(
                row=work_minus_absence_row,
                column=first_average_column + 6,
                value=avg_work_minus_absence_formula,
            )
            avg_overtime_minus_absence_cell = sheet.cell(
                row=overtime_minus_absence_row,
                column=first_average_column + 7,
                value=avg_overtime_minus_absence_formula,
            )

            avg_arrival_cell.number_format = "hh:mm"
            avg_leave_cell.number_format = "hh:mm"
            avg_work_cell.number_format = "[h]:mm"
            avg_delta_cell.number_format = "@"
            avg_overtime_cell.number_format = "[h]:mm"
            avg_absence_cell.number_format = "[h]:mm"
            avg_work_minus_absence_cell.number_format = "[h]:mm"
            avg_overtime_minus_absence_cell.number_format = "[h]:mm"


def _apply_layout(sheet: Worksheet, all_dates: List[date]) -> None:
    """Apply basic workbook style, widths and frozen panes."""
    total_columns = 5 + len(all_dates) + 9
    last_row = max(1, sheet.max_row)

    sheet.freeze_panes = "G4"
    sheet.sheet_properties.outlinePr.summaryBelow = True
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 28
    sheet.column_dimensions["E"].width = 28
    sheet.column_dimensions["F"].width = 60

    for column_index in range(7, total_columns + 1):
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = 14

    for column_index in range(1, total_columns + 1):
        header_cell = sheet.cell(row=1, column=column_index)
        header_cell.font = Font(bold=True)
        header_cell.fill = HEADER_FILL
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index in range(1, last_row + 1):
        for column_index in range(1, total_columns + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.border = THIN_BORDER
            # Skip header rows (row=1..3) to avoid overriding wrap/alignment styling.
            if row_index > 3 and column_index >= 4:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Apply header styling to date columns for header rows 2 and 3.
    if all_dates:
        first_date_column = 7
        last_date_column = first_date_column + len(all_dates) - 1
        for column_index in range(first_date_column, last_date_column + 1):
            header_cell = sheet.cell(row=2, column=column_index)
            header_cell.font = Font(bold=True)
            header_cell.fill = HEADER_FILL
            header_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for column_index in range(first_date_column, last_date_column + 1):
            header_cell = sheet.cell(row=3, column=column_index)
            header_cell.font = Font(bold=True)
            header_cell.fill = HEADER_FILL
            header_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    # Толстые рамки вокруг блока строк одного сотрудника.
    # Определяем начало блока по строкам, где в колонке F стоит "Время прихода".
    if last_row > 1:
        employee_starts = []
        for row_index in range(2, last_row + 1):
            marker = sheet.cell(row=row_index, column=6).value
            if marker == "Время прихода":
                employee_starts.append(row_index)

        for idx, start_row in enumerate(employee_starts):
            end_row = (
                employee_starts[idx + 1] - 1
                if idx + 1 < len(employee_starts)
                else last_row
            )

            # Скрываем/группируем три строки показателей для быстрого переключения в Excel (Outline).
            # Отсутствие / Длительность факт (с учетом отсутствия) / Переработка факт (с учетом отсутствия)
            for row_to_group in (start_row + 5, start_row + 6, start_row + 7):
                if row_to_group <= last_row:
                    sheet.row_dimensions[row_to_group].outlineLevel = 1
                    sheet.row_dimensions[row_to_group].hidden = True

            for column_index in range(1, total_columns + 1):
                cell = sheet.cell(row=start_row, column=column_index)
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=THICK_SIDE,
                    bottom=cell.border.bottom,
                )

                cell = sheet.cell(row=end_row, column=column_index)
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=THICK_SIDE,
                )

            for row_index in range(start_row, end_row + 1):
                cell = sheet.cell(row=row_index, column=1)
                cell.border = Border(
                    left=THICK_SIDE,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                )

                cell = sheet.cell(row=row_index, column=total_columns)
                cell.border = Border(
                    left=cell.border.left,
                    right=THICK_SIDE,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                )

    _apply_conditional_formatting(sheet, all_dates, last_row)

    # Grey-out all cells in date columns marked as "Выходной" (row >= 4),
    # so it is visually obvious that weekend days should be ignored.
    if all_dates and last_row >= 4:
        first_date_column = 7
        for day_index in range(len(all_dates)):
            col_index = first_date_column + day_index
            col_letter = get_column_letter(col_index)
            weekend_range = f"{col_letter}4:{col_letter}{last_row}"
            weekend_grey_rule = FormulaRule(
                formula=[f'{col_letter}$2="Выходной"'],
                fill=WEEKEND_GREY_FILL,
            )
            sheet.conditional_formatting.add(weekend_range, weekend_grey_rule)


def _apply_conditional_formatting(
    sheet: Worksheet, all_dates: List[date], last_row: int
) -> None:
    """Apply conditional formatting for delta and overtime rows only."""
    if not all_dates:
        return
    first_data_column = 7
    first_average_column = first_data_column + len(all_dates)
    start_col = get_column_letter(first_data_column)
    end_col = get_column_letter(first_data_column + len(all_dates) - 1)
    day_type_cell = f"{start_col}$2"
    short_day_cell = f"{start_col}$3"

    # Определяем начало блоков сотрудников по строкам с "Время прихода".
    employee_starts = []
    for row_index in range(2, last_row + 1):
        marker = sheet.cell(row=row_index, column=6).value
        if marker == "Время прихода":
            employee_starts.append(row_index)

    for employee_index, arrival_row in enumerate(employee_starts):
        delta_row = arrival_row + 3
        overtime_row = arrival_row + 4

        # Отклонение по времени прихода
        delta_range = f"{start_col}{delta_row}:{end_col}{delta_row}"
        # diff = фактическое время прихода - плановое время
        # зелёный: diff < -15 минут
        delta_green_rule = FormulaRule(
            formula=[
                f'AND({day_type_cell}="Будний",'
                f'{start_col}{arrival_row}<>"",'
                f'{start_col}{arrival_row}-$B${arrival_row}<-1/96)'
            ],
            fill=LIGHT_GREEN_FILL,
            font=BLACK_FONT,
        )
        # красный: diff > 15 минут
        delta_red_rule = FormulaRule(
            formula=[
                f'AND({day_type_cell}="Будний",'
                f'{start_col}{arrival_row}<>"",'
                f'{start_col}{arrival_row}-$B${arrival_row}>1/96)'
            ],
            fill=LIGHT_RED_FILL,
            font=BLACK_FONT,
        )
        sheet.conditional_formatting.add(delta_range, delta_green_rule)
        sheet.conditional_formatting.add(delta_range, delta_red_rule)

        # Переработка
        work_row = arrival_row + 2
        overtime_range = f"{start_col}{overtime_row}:{end_col}{overtime_row}"
        # diff = фактическая длительность - плановая длительность
        plan_duration_expr = (
            f'IF({short_day_cell}="Сокращенный",'
            f'$D${arrival_row}-TIME(1,15,0),'
            f'$D${arrival_row})'
        )
        diff_expr = f"{start_col}{work_row}-({plan_duration_expr})"

        # 0 < diff <= 30 минут  -> светло-зелёный (игнорируем почти нулевые значения)
        overtime_light_green = FormulaRule(
            formula=[
                f'AND({day_type_cell}="Будний",'
                f'{start_col}{work_row}<>"",'
                f'{diff_expr}>0,'
                f'{diff_expr}<=1/48,'
                f'ABS({diff_expr})>=1/1440)'
            ],
            fill=LIGHT_GREEN_FILL,
            font=BLACK_FONT,
        )
        # diff > 30 минут -> тёмно-зелёный
        overtime_dark_green = FormulaRule(
            formula=[
                f'AND({day_type_cell}="Будний",'
                f'{start_col}{work_row}<>"",'
                f'{diff_expr}>1/48,'
                f'ABS({diff_expr})>=1/1440)'
            ],
            fill=DARK_GREEN_FILL,
            font=BLACK_FONT,
        )
        # -30 минут <= diff < 0 -> светло-красный
        overtime_light_red = FormulaRule(
            formula=[
                f'AND({day_type_cell}="Будний",'
                f'{start_col}{work_row}<>"",'
                f'{diff_expr}<0,'
                f'{diff_expr}>=-1/48,'
                f'ABS({diff_expr})>=1/1440)'
            ],
            fill=LIGHT_RED_FILL,
            font=BLACK_FONT,
        )
        # diff < -30 минут -> тёмно-красный
        overtime_dark_red = FormulaRule(
            formula=[
                f'AND({day_type_cell}="Будний",'
                f'{start_col}{work_row}<>"",'
                f'{diff_expr}<-1/48,'
                f'ABS({diff_expr})>=1/1440)'
            ],
            fill=DARK_RED_FILL,
            font=BLACK_FONT,
        )

        sheet.conditional_formatting.add(overtime_range, overtime_light_green)
        sheet.conditional_formatting.add(overtime_range, overtime_dark_green)
        sheet.conditional_formatting.add(overtime_range, overtime_light_red)
        sheet.conditional_formatting.add(overtime_range, overtime_dark_red)

        # Переработка факт (с учетом отсутствия в течение дня)
        work_minus_absence_row = arrival_row + 6
        overtime_minus_absence_row = arrival_row + 7
        overtime_factual_range = (
            f"{start_col}{overtime_minus_absence_row}:{end_col}{overtime_minus_absence_row}"
        )
        diff_factual_expr = (
            f"{start_col}{work_minus_absence_row}-({plan_duration_expr})"
        )

        overtime_factual_light_green = FormulaRule(
            formula=[
                f'AND({day_type_cell}="Будний",'
                f'{start_col}{work_minus_absence_row}<>"",'
                f'{diff_factual_expr}>0,'
                f'{diff_factual_expr}<=1/48,'
                f'ABS({diff_factual_expr})>=1/1440)'
            ],
            fill=LIGHT_GREEN_FILL,
            font=BLACK_FONT,
        )
        overtime_factual_dark_green = FormulaRule(
            formula=[
                f'AND({day_type_cell}="Будний",'
                f'{start_col}{work_minus_absence_row}<>"",'
                f'{diff_factual_expr}>1/48,'
                f'ABS({diff_factual_expr})>=1/1440)'
            ],
            fill=DARK_GREEN_FILL,
            font=BLACK_FONT,
        )
        overtime_factual_light_red = FormulaRule(
            formula=[
                f'AND({day_type_cell}="Будний",'
                f'{start_col}{work_minus_absence_row}<>"",'
                f'{diff_factual_expr}<0,'
                f'{diff_factual_expr}>=-1/48,'
                f'ABS({diff_factual_expr})>=1/1440)'
            ],
            fill=LIGHT_RED_FILL,
            font=BLACK_FONT,
        )
        overtime_factual_dark_red = FormulaRule(
            formula=[
                f'AND({day_type_cell}="Будний",'
                f'{start_col}{work_minus_absence_row}<>"",'
                f'{diff_factual_expr}<-1/48,'
                f'ABS({diff_factual_expr})>=1/1440)'
            ],
            fill=DARK_RED_FILL,
            font=BLACK_FONT,
        )

        sheet.conditional_formatting.add(
            overtime_factual_range, overtime_factual_light_green
        )
        sheet.conditional_formatting.add(
            overtime_factual_range, overtime_factual_dark_green
        )
        sheet.conditional_formatting.add(
            overtime_factual_range, overtime_factual_light_red
        )
        sheet.conditional_formatting.add(
            overtime_factual_range, overtime_factual_dark_red
        )

    # "Среднее время работы" (avg duration) column: highlight low averages.
    # Apply across the whole avg column, but trigger only on "Длительность факт" rows.
    avg_work_col = get_column_letter(first_average_column + 2)
    avg_work_range = f"{avg_work_col}4:{avg_work_col}{last_row}"

    # 08:00 < value < 09:00 -> pale red
    avg_work_light_red = FormulaRule(
        formula=[
            f'AND($F4="Длительность факт",'
            f'{avg_work_col}4<>"",'
            f'{avg_work_col}4>TIME(8,0,0),'
            f'{avg_work_col}4<TIME(9,0,0))'
        ],
        fill=LIGHT_RED_FILL,
        font=BLACK_FONT,
    )
    # value < 08:00 -> bright red
    avg_work_bright_red = FormulaRule(
        formula=[
            f'AND($F4="Длительность факт",'
            f'{avg_work_col}4<>"",'
            f'{avg_work_col}4<TIME(8,0,0))'
        ],
        fill=DARK_RED_FILL,
        font=BLACK_FONT,
    )

    sheet.conditional_formatting.add(avg_work_range, avg_work_light_red)
    sheet.conditional_formatting.add(avg_work_range, avg_work_bright_red)